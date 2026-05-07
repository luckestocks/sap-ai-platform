"""
pages/6_War_Room.py
SAP AI Platform — Cutover Issue Tracker (War Room Mode)

A shared real-time issue board for the whole cutover team.
Any team member can log, claim, and resolve any type of issue —
data quality, load errors, config gaps, interface failures, blockers.

Design:
  - Kanban-style: Open | In Progress | Resolved columns
  - Auto-refreshes every 30 seconds
  - Any issue type, any stream — not limited to Error Analyzer outputs
  - Time tracking: raised → claimed → resolved (automatic)
  - Export full log as Excel at end of cutover
"""

import streamlit as st
import pandas as pd
from datetime import datetime, timezone, timedelta
from utils.supabase_client import get_supabase, get_clients, get_projects

st.set_page_config(
    page_title="War Room | SAP AI Platform",
    page_icon="🚨",
    layout="wide",
)

with open("components/styles.css") as f:
    st.markdown(f"<style>{f.read()}</style>", unsafe_allow_html=True)


# ── Constants ─────────────────────────────────────────────────────────────────

ISSUE_TYPES = [
    "Data Quality", "IDoc", "LTMC", "LSMW", "BAPI",
    "BODS", "RFC", "Config", "Interface", "Reconciliation", "Other"
]
STREAMS = [
    "MM", "WM", "Finance", "Procurement",
    "Manufacturing", "SOM", "Cross-Stream", "Basis"
]
PRIORITIES = ["P1 Critical", "P2 High", "P3 Medium"]
STATUSES   = ["Open", "In Progress", "Resolved", "Blocked"]
PHASES     = ["Mock", "Dress Rehearsal", "Cutover"]

PRIORITY_STYLE = {
    "P1 Critical": ("#7f1d1d", "#fca5a5", "🔴"),
    "P2 High":     ("#7c2d12", "#fdba74", "🟠"),
    "P3 Medium":   ("#713f12", "#fde68a", "🟡"),
}
STATUS_STYLE = {
    "Open":        ("#1e3a5f", "#93c5fd", "📋"),
    "In Progress": ("#1a2e1a", "#4ade80", "⚙️"),
    "Resolved":    ("#1a2a1a", "#86efac", "✅"),
    "Blocked":     ("#3b1111", "#f87171", "🚫"),
}


# ── Helpers ───────────────────────────────────────────────────────────────────

def now_utc() -> str:
    return datetime.now(timezone.utc).isoformat()


def fmt_duration(start_iso: str, end_iso: str = None) -> str:
    """Return human-readable duration between two ISO timestamps."""
    try:
        start = datetime.fromisoformat(start_iso.replace("Z", "+00:00"))
        end   = datetime.fromisoformat(end_iso.replace("Z", "+00:00")) if end_iso else datetime.now(timezone.utc)
        delta = end - start
        total = int(delta.total_seconds())
        if total < 60:
            return f"{total}s"
        if total < 3600:
            return f"{total // 60}m"
        h, m = divmod(total // 60, 60)
        return f"{h}h {m}m"
    except Exception:
        return "—"


def load_issues(project_id: str = None) -> list[dict]:
    """Load all war room issues, optionally filtered by project."""
    try:
        supabase = get_supabase()
        q = supabase.table("war_room_issues").select("*").order("raised_at", desc=False)
        if project_id:
            q = q.eq("project_id", project_id)
        res = q.execute()
        return res.data or []
    except Exception as e:
        st.error(f"Failed to load issues: {e}")
        return []


def create_issue(payload: dict) -> bool:
    try:
        get_supabase().table("war_room_issues").insert(payload).execute()
        return True
    except Exception as e:
        st.error(f"Failed to create issue: {e}")
        return False


def update_issue(issue_id: str, updates: dict) -> bool:
    try:
        get_supabase().table("war_room_issues").update(updates).eq("id", issue_id).execute()
        return True
    except Exception as e:
        st.error(f"Failed to update issue: {e}")
        return False


# ── Priority / status pills (HTML) ───────────────────────────────────────────

def priority_pill(priority: str) -> str:
    bg, fg, icon = PRIORITY_STYLE.get(priority, ("#1e293b", "#94a3b8", "⚪"))
    return (
        f'<span style="background:{bg};color:{fg};border-radius:10px;'
        f'padding:2px 8px;font-size:0.70rem;font-weight:700;">{icon} {priority}</span>'
    )


def status_pill(status: str) -> str:
    bg, fg, icon = STATUS_STYLE.get(status, ("#1e293b", "#94a3b8", "•"))
    return (
        f'<span style="background:{bg};color:{fg};border-radius:10px;'
        f'padding:2px 8px;font-size:0.70rem;font-weight:700;">{icon} {status}</span>'
    )


def type_pill(issue_type: str) -> str:
    return (
        f'<span style="background:#1e3a5f;color:#93c5fd;border-radius:6px;'
        f'padding:2px 8px;font-size:0.70rem;font-weight:600;">{issue_type}</span>'
    )


def stream_pill(stream: str) -> str:
    return (
        f'<span style="background:#1e293b;color:#64748b;border-radius:6px;'
        f'padding:2px 8px;font-size:0.68rem;">{stream}</span>'
    )


# ── Issue card renderer ───────────────────────────────────────────────────────

def render_issue_card(issue: dict, your_name: str):
    """Render a single issue card with action buttons."""
    iid      = issue["id"]
    status   = issue["status"]
    priority = issue["priority"]
    bg_map   = {
        "P1 Critical": "#1a0505",
        "P2 High":     "#1a0d00",
        "P3 Medium":   "#1a1200",
    }
    border_map = {
        "P1 Critical": "#ef4444",
        "P2 High":     "#f97316",
        "P3 Medium":   "#eab308",
    }
    bg     = bg_map.get(priority, "#0f172a")
    border = border_map.get(priority, "#334155")

    # Time open
    time_open = fmt_duration(issue["raised_at"])
    time_to_resolve = ""
    if issue.get("resolved_at") and issue.get("raised_at"):
        time_to_resolve = fmt_duration(issue["raised_at"], issue["resolved_at"])

    # Split description from embedded screenshot marker
    raw_desc = issue.get("description") or ""
    screenshot_data = None
    display_desc    = raw_desc
    if "[SCREENSHOT:" in raw_desc:
        parts        = raw_desc.split("[SCREENSHOT:", 1)
        display_desc = parts[0].strip()
        try:
            screenshot_data = parts[1].rstrip("]")
        except Exception:
            screenshot_data = None

    st.markdown(
        f'<div style="background:{bg};border:1px solid {border};border-radius:8px;'
        f'padding:10px 12px;margin-bottom:8px;">'
        f'<div style="display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:6px;">'
        f'<div style="display:flex;gap:4px;flex-wrap:wrap;">'
        f'{priority_pill(priority)}{type_pill(issue["issue_type"])}{stream_pill(issue["stream"])}'
        f'</div>'
        f'<span style="color:#475569;font-size:0.68rem;">⏱ {time_open}</span>'
        f'</div>'
        f'<div style="color:#e2e8f0;font-size:0.85rem;font-weight:600;margin-bottom:4px;">'
        f'{issue["title"]}</div>'
        + (f'<div style="color:#94a3b8;font-size:0.75rem;margin-bottom:6px;">{display_desc}</div>'
           if display_desc else "")
        + f'<div style="display:flex;justify-content:space-between;align-items:center;">'
        f'<span style="color:#475569;font-size:0.70rem;">🙋 {issue["raised_by"]}'
        + (f' → ⚙️ {issue["claimed_by"]}' if issue.get("claimed_by") else "")
        + (f' → ✅ {issue["resolved_by"]} ({time_to_resolve})' if issue.get("resolved_by") else "")
        + f'</span></div>'
        + (f'<div style="color:#4ade80;font-size:0.72rem;margin-top:4px;border-top:1px solid #1e293b;padding-top:4px;">✅ {issue["resolution_notes"]}</div>'
           if issue.get("resolution_notes") else "")
        + f'</div>',
        unsafe_allow_html=True,
    )

    # Show screenshot if attached
    if screenshot_data:
        with st.expander("📸 Screenshot", expanded=False):
            st.markdown(f'<img src="{screenshot_data}" style="max-width:100%;border-radius:6px;">',
                        unsafe_allow_html=True)

    # Action buttons — Resolved issues have no actions
    if status == "Resolved":
        return

    btn_col1, btn_col2, btn_col3 = st.columns([1, 1, 1])

    if status == "Open":
        with btn_col1:
            if st.button("⚙️ Claim", key=f"claim_{iid}", use_container_width=True):
                update_issue(iid, {
                    "status":     "In Progress",
                    "claimed_by": your_name,
                    "claimed_at": now_utc(),
                })
                st.rerun()
        with btn_col2:
            if st.button("🚫 Block", key=f"block_{iid}", use_container_width=True):
                update_issue(iid, {"status": "Blocked"})
                st.rerun()

    elif status == "In Progress":
        with btn_col1:
            if st.button("✅ Resolve", key=f"resolve_{iid}", use_container_width=True):
                st.session_state["resolve_panel_id"] = iid
                st.session_state["resolve_panel_title"] = issue["title"]
        with btn_col2:
            if st.button("🚫 Block", key=f"block2_{iid}", use_container_width=True):
                update_issue(iid, {"status": "Blocked"})
                st.rerun()

    elif status == "Blocked":
        with btn_col1:
            if st.button("↩️ Reopen", key=f"reopen_{iid}", use_container_width=True):
                update_issue(iid, {"status": "Open", "claimed_by": None, "claimed_at": None})
                st.rerun()
        with btn_col2:
            if st.button("⚙️ Claim", key=f"claim2_{iid}", use_container_width=True):
                update_issue(iid, {
                    "status":     "In Progress",
                    "claimed_by": your_name,
                    "claimed_at": now_utc(),
                })
                st.rerun()


# ── Page header ───────────────────────────────────────────────────────────────

st.markdown("# 🚨 War Room")
st.markdown("Cutover Issue Tracker — log, claim, and resolve issues in real time.")

# ── Your name (persisted in session) ─────────────────────────────────────────
if "war_room_name" not in st.session_state:
    st.session_state["war_room_name"] = ""

if not st.session_state["war_room_name"]:
    st.info("👤 Enter your name to start using the War Room.")
    name_col, btn_col = st.columns([3, 1])
    with name_col:
        entered_name = st.text_input("Your name", placeholder="e.g. Sparky", label_visibility="collapsed")
    with btn_col:
        if st.button("Join War Room", type="primary", use_container_width=True) and entered_name.strip():
            st.session_state["war_room_name"] = entered_name.strip()
            st.rerun()
    st.stop()

your_name = st.session_state["war_room_name"]

# ── Top bar: name + project context + phase + refresh ────────────────────────
top1, top2, top3, top4, top5 = st.columns([2, 2, 2, 1, 1])

with top1:
    st.markdown(
        f'<div style="background:#1a2e1a;border:1px solid #4ade80;border-radius:8px;'
        f'padding:6px 12px;font-size:0.82rem;color:#4ade80;font-weight:600;">'
        f'👤 {your_name}</div>',
        unsafe_allow_html=True,
    )
    if st.button("Change name", key="change_name"):
        st.session_state["war_room_name"] = ""
        st.rerun()

with top2:
    clients    = get_clients()
    client_map = {c["name"]: c["id"] for c in clients}
    sel_client = st.selectbox("Client", ["— all —"] + list(client_map.keys()), key="wr_client")
    wr_client_id = client_map.get(sel_client)

with top3:
    wr_project_id = None
    if wr_client_id:
        from utils.supabase_client import get_projects
        projects    = get_projects(wr_client_id)
        proj_map    = {p["name"]: p["id"] for p in projects}
        sel_project = st.selectbox("Project", ["— all —"] + list(proj_map.keys()), key="wr_project")
        wr_project_id = proj_map.get(sel_project)
    else:
        st.selectbox("Project", ["— select client —"], disabled=True, key="wr_proj_dis")

with top4:
    cutover_phase = st.selectbox("Phase", PHASES, index=2, key="wr_phase")

with top5:
    auto_refresh = st.checkbox("🔄 Auto", value=True, help="Auto-refresh every 30s")

st.divider()

# ── Load all issues ───────────────────────────────────────────────────────────
all_issues = load_issues(project_id=wr_project_id)

open_issues     = [i for i in all_issues if i["status"] == "Open"]
inprog_issues   = [i for i in all_issues if i["status"] == "In Progress"]
resolved_issues = [i for i in all_issues if i["status"] == "Resolved"]
blocked_issues  = [i for i in all_issues if i["status"] == "Blocked"]

# ── Summary metrics ───────────────────────────────────────────────────────────
m1, m2, m3, m4, m5 = st.columns(5)
m1.metric("📋 Open",        len(open_issues))
m2.metric("⚙️ In Progress", len(inprog_issues))
m3.metric("🚫 Blocked",     len(blocked_issues))
m4.metric("✅ Resolved",    len(resolved_issues))
m5.metric("📊 Total",       len(all_issues))

st.divider()

# ── Log new issue ─────────────────────────────────────────────────────────────
with st.expander("➕ Log New Issue", expanded=len(all_issues) == 0):

    # Paste screenshot — must live OUTSIDE st.form()
    from streamlit_paste_button import paste_image_button as pbutton
    import io as _io
    import base64

    st.caption("📋 Attach a screenshot (optional) — Win+Shift+S to snip, then click below and Ctrl+V")
    paste_result = pbutton(
        label="📋 Paste Screenshot",
        background_color="#1E3A5F",
        hover_background_color="#1d6fa5",
        key="wr_paste_screenshot",
    )
    if paste_result.image_data is not None:
        st.session_state["wr_screenshot"] = paste_result.image_data
    if st.session_state.get("wr_screenshot"):
        st.image(st.session_state["wr_screenshot"], caption="Screenshot attached", width=400)
        if st.button("✖ Remove screenshot", key="wr_remove_screenshot"):
            st.session_state.pop("wr_screenshot", None)
            st.rerun()

    with st.form("new_issue_form"):
        ni1, ni2, ni3 = st.columns(3)
        with ni1:
            ni_priority = st.selectbox("Priority *", PRIORITIES, key="ni_priority")
        with ni2:
            ni_type = st.selectbox("Issue Type *", ISSUE_TYPES, key="ni_type")
        with ni3:
            ni_stream = st.selectbox("Stream *", STREAMS, key="ni_stream")

        ni_title = st.text_input(
            "Title *",
            placeholder="e.g. Vendor master missing plant data — 1200 records affected",
        )
        ni_desc = st.text_area(
            "Description (optional)",
            height=80,
            placeholder="Additional context, error codes, transaction, object count...",
        )

        submit_issue = st.form_submit_button("🚨 Log Issue", type="primary")

    if submit_issue:
        if not ni_title.strip():
            st.warning("Please add a title.")
        else:
            # Convert screenshot to base64 and append to description if present
            desc_text = ni_desc.strip()
            if st.session_state.get("wr_screenshot"):
                try:
                    buf = _io.BytesIO()
                    st.session_state["wr_screenshot"].save(buf, format="PNG")
                    b64 = base64.b64encode(buf.getvalue()).decode()
                    # Store marker so card renderer can detect and show it
                    desc_text = (desc_text + "\n" if desc_text else "") + f"[SCREENSHOT:data:image/png;base64,{b64}]"
                except Exception:
                    pass  # screenshot encode failed — log without it

            ok = create_issue({
                "title":         ni_title.strip(),
                "description":   desc_text or None,
                "issue_type":    ni_type,
                "stream":        ni_stream,
                "priority":      ni_priority,
                "status":        "Open",
                "raised_by":     your_name,
                "raised_at":     now_utc(),
                "cutover_phase": cutover_phase,
                "client_id":     wr_client_id,
                "project_id":    wr_project_id,
            })
            if ok:
                st.session_state.pop("wr_screenshot", None)
                st.success(f"✅ Issue logged: {ni_title.strip()}")
                st.rerun()

# ── Filters ───────────────────────────────────────────────────────────────────
with st.expander("🔽 Filters", expanded=False):
    f1, f2, f3 = st.columns(3)
    with f1:
        filter_priority = st.multiselect("Priority", PRIORITIES, default=[])
    with f2:
        filter_type = st.multiselect("Issue Type", ISSUE_TYPES, default=[])
    with f3:
        filter_stream = st.multiselect("Stream", STREAMS, default=[])

def apply_filters(issues: list) -> list:
    if filter_priority:
        issues = [i for i in issues if i["priority"] in filter_priority]
    if filter_type:
        issues = [i for i in issues if i["issue_type"] in filter_type]
    if filter_stream:
        issues = [i for i in issues if i["stream"] in filter_stream]
    return issues

# ── Resolve panel — full width, above board, only shown when active ───────────
resolve_id    = st.session_state.get("resolve_panel_id")
resolve_title = st.session_state.get("resolve_panel_title", "")

if resolve_id:
    st.markdown(
        f'<div style="background:#0d2218;border:2px solid #4ade80;border-radius:10px;'
        f'padding:12px 16px;margin-bottom:16px;">'
        f'<span style="color:#4ade80;font-weight:700;font-size:0.90rem;">✅ Resolving: </span>'
        f'<span style="color:#e2e8f0;font-size:0.88rem;">{resolve_title}</span>'
        f'</div>',
        unsafe_allow_html=True,
    )
    with st.form("resolve_panel_form"):
        notes = st.text_area(
            "Resolution notes *",
            height=90,
            placeholder="Describe exactly what was done to fix this...",
        )
        rc1, rc2 = st.columns([1, 3])
        with rc1:
            cancel = st.form_submit_button("✖ Cancel")
        with rc2:
            save = st.form_submit_button("💾 Save & Close Issue", type="primary")

        if cancel:
            st.session_state.pop("resolve_panel_id", None)
            st.session_state.pop("resolve_panel_title", None)
            st.rerun()
        if save:
            if not notes.strip():
                st.warning("Please add resolution notes before closing.")
            else:
                update_issue(resolve_id, {
                    "status":           "Resolved",
                    "resolved_by":      your_name,
                    "resolved_at":      now_utc(),
                    "resolution_notes": notes.strip(),
                })
                st.session_state.pop("resolve_panel_id", None)
                st.session_state.pop("resolve_panel_title", None)
                st.rerun()

# ── Kanban board ──────────────────────────────────────────────────────────────
# Sort: P1 first, then P2, then P3; within priority by raised_at
PRIORITY_ORDER = {"P1 Critical": 0, "P2 High": 1, "P3 Medium": 2}

def sort_issues(issues: list) -> list:
    return sorted(issues, key=lambda i: (PRIORITY_ORDER.get(i["priority"], 9), i["raised_at"]))

col_open, col_inprog, col_resolved = st.columns(3)

with col_open:
    filtered_open = apply_filters(sort_issues(open_issues))
    st.markdown(
        f'<div style="background:#1e3a5f;color:#93c5fd;border-radius:8px;'
        f'padding:6px 12px;font-weight:700;font-size:0.85rem;margin-bottom:12px;">'
        f'📋 Open &nbsp;<span style="background:#0f172a;border-radius:10px;'
        f'padding:1px 8px;font-size:0.75rem;">{len(filtered_open)}</span></div>',
        unsafe_allow_html=True,
    )
    if not filtered_open:
        st.caption("No open issues.")
    for issue in filtered_open:
        render_issue_card(issue, your_name)

with col_inprog:
    filtered_inprog = apply_filters(sort_issues(inprog_issues + blocked_issues))
    st.markdown(
        f'<div style="background:#1a2e1a;color:#4ade80;border-radius:8px;'
        f'padding:6px 12px;font-weight:700;font-size:0.85rem;margin-bottom:12px;">'
        f'⚙️ In Progress &nbsp;<span style="background:#0f172a;border-radius:10px;'
        f'padding:1px 8px;font-size:0.75rem;">{len(filtered_inprog)}</span></div>',
        unsafe_allow_html=True,
    )
    if not filtered_inprog:
        st.caption("No issues in progress.")
    for issue in filtered_inprog:
        render_issue_card(issue, your_name)

with col_resolved:
    filtered_resolved = apply_filters(sort_issues(resolved_issues))
    # Show most recently resolved first
    filtered_resolved = sorted(filtered_resolved,
                                key=lambda i: i.get("resolved_at") or "", reverse=True)
    st.markdown(
        f'<div style="background:#1a2a1a;color:#86efac;border-radius:8px;'
        f'padding:6px 12px;font-weight:700;font-size:0.85rem;margin-bottom:12px;">'
        f'✅ Resolved &nbsp;<span style="background:#0f172a;border-radius:10px;'
        f'padding:1px 8px;font-size:0.75rem;">{len(filtered_resolved)}</span></div>',
        unsafe_allow_html=True,
    )
    if not filtered_resolved:
        st.caption("No resolved issues yet.")
    for issue in filtered_resolved:
        render_issue_card(issue, your_name)

# ── Export ────────────────────────────────────────────────────────────────────
st.divider()

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def build_export_excel(issues: list) -> bytes:
    """Build the War Room Excel export. Returns bytes ready for download."""
    wb  = Workbook()
    ws1 = wb.active
    ws1.title = "Issue Log"

    headers = [
        "Issue #", "Priority", "Status", "Type", "Stream",
        "Title", "Description", "Raised By", "Raised At",
        "Claimed By", "Claimed At", "Resolved By", "Resolved At",
        "Time to Resolve", "Resolution Notes", "Phase"
    ]
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="93C5FD", size=10)
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    priority_fills = {
        "P1 Critical": PatternFill("solid", fgColor="7F1D1D"),
        "P2 High":     PatternFill("solid", fgColor="7C2D12"),
        "P3 Medium":   PatternFill("solid", fgColor="713F12"),
    }

    def fmt_ts(ts):
        if not ts:
            return ""
        try:
            dt = datetime.fromisoformat(ts.replace("Z", "+00:00"))
            return dt.strftime("%Y-%m-%d %H:%M")
        except Exception:
            return ts

    sorted_issues = sort_issues(issues)
    for row_num, issue in enumerate(sorted_issues, 2):
        ttr = fmt_duration(issue["raised_at"], issue["resolved_at"]) if issue.get("resolved_at") else ""
        row_data = [
            row_num - 1,
            issue["priority"],
            issue["status"],
            issue["issue_type"],
            issue["stream"],
            issue["title"],
            issue.get("description") or "",
            issue["raised_by"],
            fmt_ts(issue["raised_at"]),
            issue.get("claimed_by") or "",
            fmt_ts(issue.get("claimed_at")),
            issue.get("resolved_by") or "",
            fmt_ts(issue.get("resolved_at")),
            ttr,
            issue.get("resolution_notes") or "",
            issue.get("cutover_phase") or "",
        ]
        fill = priority_fills.get(issue["priority"], PatternFill("solid", fgColor="0F172A"))
        for col, val in enumerate(row_data, 1):
            cell = ws1.cell(row=row_num, column=col, value=val)
            cell.font = Font(color="E2E8F0", size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.fill = fill

    col_widths = [8, 14, 14, 14, 14, 40, 35, 14, 18, 14, 18, 14, 18, 14, 40, 16]
    for i, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.row_dimensions[1].height = 20
    ws1.freeze_panes = "A2"

    # Sheet 2: Summary
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "War Room Summary"
    ws2["A1"].font = Font(bold=True, size=14, color="93C5FD")

    summary_rows = [
        ("", ""),
        ("Total Issues",  len(issues)),
        ("Open",          sum(1 for i in issues if i["status"] == "Open")),
        ("In Progress",   sum(1 for i in issues if i["status"] == "In Progress")),
        ("Blocked",       sum(1 for i in issues if i["status"] == "Blocked")),
        ("Resolved",      sum(1 for i in issues if i["status"] == "Resolved")),
        ("", ""),
        ("By Priority", ""),
        ("P1 Critical",   sum(1 for i in issues if i["priority"] == "P1 Critical")),
        ("P2 High",       sum(1 for i in issues if i["priority"] == "P2 High")),
        ("P3 Medium",     sum(1 for i in issues if i["priority"] == "P3 Medium")),
        ("", ""),
        ("By Stream", ""),
    ]
    for stream in STREAMS:
        count = sum(1 for i in issues if i["stream"] == stream)
        if count:
            summary_rows.append((stream, count))

    for r, (label, value) in enumerate(summary_rows, 3):
        ws2.cell(row=r, column=1, value=label).font = Font(color="94A3B8", size=10, bold=(value == ""))
        if value != "":
            ws2.cell(row=r, column=2, value=value).font = Font(color="E2E8F0", size=10, bold=True)

    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


exp_col1, exp_col2 = st.columns([3, 1])
with exp_col1:
    st.markdown("### 📥 Export War Room Log")
    st.caption("Download the full issue log as Excel — use this as your post-cutover resolution report.")
with exp_col2:
    if not all_issues:
        st.button("📥 Export Excel", disabled=True, use_container_width=True,
                  help="No issues to export yet.")
    else:
        phase_tag = cutover_phase.replace(" ", "_")
        filename  = f"WarRoom_{phase_tag}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        st.download_button(
            label="📥 Export Excel",
            data=build_export_excel(all_issues),
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary",
        )

# ── Auto-refresh ──────────────────────────────────────────────────────────────
if auto_refresh:
    import time
    time.sleep(30)
    st.rerun()
